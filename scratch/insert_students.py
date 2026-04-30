import openpyxl
import urllib.request
import urllib.error
import json
import sys
import ssl

# Mac Python 3.14 SSL cert fix
ssl_ctx = ssl.create_default_context()
ssl_ctx.check_hostname = False
ssl_ctx.verify_mode = ssl.CERT_NONE

SUPABASE_URL = "https://ayevvaecybqjvlvmrbme.supabase.co"
SERVICE_KEY  = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImF5ZXZ2YWVjeWJxanZsdm1yYm1lIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NzQ1MTg1MiwiZXhwIjoyMDkzMDI3ODUyfQ.mIyomUW0MtEvGtUVYaKmbiIdHPW-4For-6b0YRtfCjg"
EXCEL_PATH   = "scratch/students.xlsx"

# ── 1. Parse Excel ────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL_PATH)
students = []

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n📄 Sheet: {sheet_name}")
    found_header = False
    for row in ws.iter_rows(values_only=True):
        # Look for the header row containing "HT No" or "HT NO"
        if not found_header:
            row_str = " ".join(str(c).upper() for c in row if c)
            if "HT" in row_str and ("NO" in row_str or "NUMBER" in row_str):
                found_header = True
                # Detect which columns are HT No and Name
                headers = [str(c).strip().upper() if c else "" for c in row]
                ht_col = next((i for i, h in enumerate(headers) if "HT" in h), None)
                name_col = next((i for i, h in enumerate(headers) if "NAME" in h or "STUDENT" in h), None)
                print(f"   Found header → HT col={ht_col}, Name col={name_col}")
            continue

        if ht_col is None or name_col is None:
            continue

        ht   = row[ht_col]  if ht_col  < len(row) else None
        name = row[name_col] if name_col < len(row) else None

        # Skip blank / section header rows
        if not ht or not name:
            continue
        ht_str   = str(ht).strip()
        name_str = str(name).strip()
        if not ht_str or not name_str or len(ht_str) < 5:
            continue
        # Skip if HT looks like a header again
        if "HT" in ht_str.upper() and "NO" in ht_str.upper():
            continue

        students.append({"name": name_str, "unique_id": ht_str, "role": "student"})
        print(f"   ✅ {ht_str} — {name_str}")

print(f"\n📊 Total students parsed: {len(students)}")
if not students:
    print("❌ No students found — check column headers in the sheet.")
    sys.exit(1)

# ── 2. Insert into Supabase (upsert to avoid duplicates) ─────────────────────
url  = f"{SUPABASE_URL}/rest/v1/qr_users"
data = json.dumps(students).encode("utf-8")

req = urllib.request.Request(
    url,
    data=data,
    headers={
        "apikey":        SERVICE_KEY,
        "Authorization": f"Bearer {SERVICE_KEY}",
        "Content-Type":  "application/json",
        "Prefer":        "resolution=merge-duplicates",
        "on-conflict":   "unique_id"
    },
    method="POST",
)

try:
    with urllib.request.urlopen(req, context=ssl_ctx) as resp:
        body = resp.read().decode()
        print(f"\n🎉 Inserted {len(students)} students successfully!")
        print(f"   Status: {resp.status}")
except urllib.error.HTTPError as e:
    body = e.read().decode()
    print(f"\n❌ HTTP {e.code}: {body}")
    sys.exit(1)
