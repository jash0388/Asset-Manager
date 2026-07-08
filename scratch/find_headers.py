import openpyxl

wb = openpyxl.load_workbook("/Users/jashwanthsingh/Downloads/DS-SEM-1 roll list-A.Y-2026-27.xlsx")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- Sheet {sheet_name} ---")
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        row_str = " ".join(str(cell) for cell in row if cell is not None)
        if "Class / Section" in row_str or "Batch" in row_str or "Admn No" in row_str or "HT No" in row_str:
            print(f"Row {r_idx+1}:", [str(c) if c is not None else "" for c in row])
