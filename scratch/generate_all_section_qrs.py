import os
import re
import shutil
import zipfile
from io import BytesIO
import openpyxl
import qrcode
from PIL import Image, ImageDraw, ImageFont
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas

EXCEL_PATH   = "/Users/jashwanthsingh/Downloads/DS-SEM-1 roll list-A.Y-2026-27.xlsx"
OUTPUT_DIR   = "student_qr_codes"
ZIP_FILE     = "SPHN_Student_QR_Codes.zip"

# Reset the output directory completely
if os.path.exists(OUTPUT_DIR):
    shutil.rmtree(OUTPUT_DIR)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# 1. Parse Excel Sheet and Group by Section
print(f"📖 Parsing Excel file: {EXCEL_PATH}")
wb = openpyxl.load_workbook(EXCEL_PATH)
sections_data = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    current_section = None
    ht_col = None
    name_col = None
    
    for row in ws.iter_rows(values_only=True):
        row_str = " ".join(str(c) for c in row if c is not None)
        
        # Check for section header
        if "Class / Section" in row_str:
            match = re.search(r"Class\s*/\s*Section:\s*([^\n\r]+)", row_str, re.IGNORECASE)
            if match:
                current_section = match.group(1).strip().replace(" ", "_").replace("/", "-")
                sections_data[current_section] = []
            continue
            
        # Check for column headers
        row_upper = [str(c).strip().upper() if c is not None else "" for c in row]
        if any("HT" in h or "ADMN" in h for h in row_upper) and any("NAME" in h or "STUDENT" in h for h in row_upper):
            ht_col = next((i for i, h in enumerate(row_upper) if "HT" in h or "ADMN" in h), None)
            name_col = next((i for i, h in enumerate(row_upper) if "NAME" in h or "STUDENT" in h), None)
            continue
            
        # Parse data row
        if current_section and ht_col is not None and name_col is not None:
            if len(row) <= max(ht_col, name_col):
                continue
            ht_val = row[ht_col]
            name_val = row[name_col]
            
            if ht_val is None or name_val is None:
                continue
                
            ht_str = str(ht_val).strip()
            name_str = str(name_val).strip()
            
            if not ht_str or not name_str or len(ht_str) < 5 or ht_str.upper() in ["HT NO", "ADMN NO"]:
                continue
                
            sections_data[current_section].append({
                "name": name_str,
                "roll_no": ht_str
            })

print("   ✓ Finished parsing all sections.")
for sec, studs in sections_data.items():
    print(f"     - Section {sec}: {len(studs)} students")

# PDF Layout Config
PAGE_W, PAGE_H = A4
COLS      = 4
MARGIN    = 10 * mm
QR_SIZE   = 42 * mm
CELL_W    = (PAGE_W - 2 * MARGIN) / COLS
CELL_H    = QR_SIZE + 16 * mm
ROWS_PER  = int((PAGE_H - 2 * MARGIN) / CELL_H)

def draw_qr_on_pdf(c, student, x, y):
    name = student["name"].strip()
    roll = student["roll_no"].strip()

    # Generate QR (encodes roll number/unique_id directly)
    qr = qrcode.QRCode(version=1, box_size=10, border=2)
    qr.add_data(roll)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white").convert("RGB")

    buf = BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    
    # Draw QR Image on PDF
    c.drawImage(ImageReader(buf), x + (CELL_W - QR_SIZE) / 2, y + 10 * mm, width=QR_SIZE, height=QR_SIZE)

    # Draw Name
    display_name = name if len(name) <= 22 else name[:20] + "…"
    c.setFont("Helvetica-Bold", 7)
    c.setFillColor(colors.black)
    c.drawCentredString(x + CELL_W / 2, y + 6 * mm, display_name)

    # Draw Roll Number
    c.setFont("Helvetica", 6.5)
    c.setFillColor(colors.grey)
    c.drawCentredString(x + CELL_W / 2, y + 2 * mm, roll)

    # Draw border card outline
    c.setStrokeColor(colors.HexColor("#dddddd"))
    c.setLineWidth(0.5)
    c.rect(x + 1*mm, y + 0.5*mm, CELL_W - 2*mm, CELL_H - 1*mm)

all_created_items = []

for section_name, students in sections_data.items():
    if not students:
        continue
    
    print(f"\n🎨 Processing Section: {section_name} ({len(students)} students)")
    sec_png_dir = os.path.join(OUTPUT_DIR, section_name)
    os.makedirs(sec_png_dir, exist_ok=True)
    
    # 1. Generate individual labeled PNG files
    for s in students:
        roll = s["roll_no"]
        name = s["name"]
        
        safe_roll = roll.replace("/", "_").replace(" ", "")
        png_path = os.path.join(sec_png_dir, f"{safe_roll}.png")
        
        qr = qrcode.QRCode(version=1, box_size=10, border=3)
        qr.add_data(roll)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
        
        # Add label margins under PNG image
        qr_w, qr_h = qr_img.size
        canvas_h = qr_h + 80
        canvas_img = Image.new("RGB", (qr_w, canvas_h), "white")
        canvas_img.paste(qr_img, (0, 0))
        
        draw = ImageDraw.Draw(canvas_img)
        try:
            font = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", 20)
            font_small = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", 16)
        except Exception:
            font = ImageFont.load_default()
            font_small = font
            
        r_box = draw.textbbox((0, 0), roll, font=font)
        draw.text(((qr_w - (r_box[2] - r_box[0])) / 2, qr_h + 5), roll, fill="black", font=font)
        
        n_box = draw.textbbox((0, 0), name, font=font_small)
        draw.text(((qr_w - (n_box[2] - n_box[0])) / 2, qr_h + 35), name, fill="black", font=font_small)
        
        canvas_img.save(png_path)
        all_created_items.append((png_path, os.path.join(section_name, f"{safe_roll}.png")))
        
    # 2. Generate section PDF file
    pdf_path = os.path.join(OUTPUT_DIR, f"{section_name}_printable.pdf")
    pdf_c = canvas.Canvas(pdf_path, pagesize=A4)
    pdf_c.setTitle(f"QR Codes - {section_name.replace('_', ' ')}")
    
    idx = 0
    total = len(students)
    while idx < total:
        page_start_y = PAGE_H - MARGIN - CELL_H
        for row in range(ROWS_PER):
            for col in range(COLS):
                if idx >= total:
                    break
                x = MARGIN + col * CELL_W
                y = page_start_y - row * CELL_H
                draw_qr_on_pdf(pdf_c, students[idx], x, y)
                idx += 1
                
        # Draw footer
        pdf_c.setFont("Helvetica", 7)
        pdf_c.setFillColor(colors.grey)
        pdf_c.drawCentredString(PAGE_W / 2, 6 * mm, f"QR Section: {section_name.replace('_', ' ')} — Page {pdf_c.getPageNumber()}")
        
        if idx < total:
            pdf_c.showPage()
            
    pdf_c.save()
    print(f"   ✓ Generated PDF sheet: {pdf_path}")
    all_created_items.append((pdf_path, f"{section_name}_printable.pdf"))

# Package everything in a ZIP file
print(f"\n📦 Packaging all 400+ student assets into ZIP file: {ZIP_FILE}")
with zipfile.ZipFile(ZIP_FILE, "w", zipfile.ZIP_DEFLATED) as zf:
    for local_path, arc_path in all_created_items:
        zf.write(local_path, arc_path)

print("🎉 ZIP package with all 400+ students and section PDFs generated successfully!")
