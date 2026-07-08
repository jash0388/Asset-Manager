import openpyxl

wb = openpyxl.load_workbook("/Users/jashwanthsingh/Downloads/DS-SEM-1 roll list-A.Y-2026-27.xlsx")
print("Sheets:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\nSheet {sheet_name} (Max row {ws.max_row}):")
    # print first 10 rows
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 10:
            break
        print(f"Row {i+1}:", [str(cell)[:30] if cell is not None else "" for cell in row])
