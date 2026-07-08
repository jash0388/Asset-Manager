import os
import re
import sys
import zipfile
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas

EXCEL_PATH   = "/Users/jashwanthsingh/Downloads/DS-SEM-1 roll list-A.Y-2026-27.xlsx"
QR_DIR       = "student_qr_codes"
OUTPUT_ZIP   = "SPHN_Student_QR_Codes.zip"

print(f"📖 Reading student name mappings from Excel: {EXCEL_PATH}")
wb = openpyxl.load_workbook(EXCEL_PATH)
roll_name_map = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    ht_col = None
    name_col = None
    
    for row in ws.iter_rows(values_only=True):
        row_str = " ".join(str(c) for c in row if c is not None)
        
        # Detect headers
        row_upper = [str(c).strip().upper() if c is not None else "" for c in row]
        if any("HT" in h or "ADMN" in h for h in row_upper) and any("NAME" in h or "STUDENT" in h for h in row_upper):
            ht_col = next((i for i, h in enumerate(row_upper) if "HT" in h or "ADMN" in h), None)
            name_col = next((i for i, h in enumerate(row_upper) if "NAME" in h or "STUDENT" in h), None)
            continue
            
        if ht_col is not None and name_col is not None:
            if len(row) <= max(ht_col, name_col):
                continue
            ht_val = row[ht_col]
            name_val = row[name_col]
            if ht_val is not None and name_val is not None:
                roll = str(ht_val).strip()
                name = str(name_val).strip()
                roll_name_map[roll] = name

print(f"   ✓ Loaded {len(roll_name_map)} student roll-name mappings.")

# PDF Layout Config
PAGE_W, PAGE_H = A4
COLS      = 4
MARGIN    = 10 * mm
QR_SIZE   = 42 * mm
CELL_W    = (PAGE_W - 2 * MARGIN) / COLS
CELL_H    = QR_SIZE + 16 * mm
ROWS_PER  = int((PAGE_H - 2 * MARGIN) / CELL_H)

def generate_pdf_for_folder(folder_name):
    folder_path = os.path.join(QR_DIR, folder_name)
    if not os.path.exists(folder_path):
        print(f"⚠️ Folder {folder_path} does not exist.")
        return None
        
    png_files = [f for f in os.listdir(folder_path) if f.lower().endswith(".png")]
    png_files.sort()
    
    pdf_path = os.path.join(QR_DIR, f"{folder_name}_printable.pdf")
    print(f"\n🎨 Generating PDF sheet for {folder_name} ({len(png_files)} QR codes) -> {pdf_path}")
    
    pdf_c = canvas.Canvas(pdf_path, pagesize=A4)
    pdf_c.setTitle(f"QR Codes - {folder_name.replace('_', ' ')}")
    
    idx = 0
    total = len(png_files)
    
    while idx < total:
        page_start_y = PAGE_H - MARGIN - CELL_H
        for row in range(ROWS_PER):
            for col in range(COLS):
                if idx >= total:
                    break
                filename = png_files[idx]
                roll = os.path.splitext(filename)[0]
                name = roll_name_map.get(roll, "Unknown Student")
                
                # Math coordinate
                x = MARGIN + col * CELL_W
                y = page_start_y - row * CELL_H
                
                # Draw border card outline
                pdf_c.setStrokeColor(colors.HexColor("#dddddd"))
                pdf_c.setLineWidth(0.5)
                pdf_c.rect(x + 1*mm, y + 0.5*mm, CELL_W - 2*mm, CELL_H - 1*mm)
                
                # Draw QR code PNG image directly
                image_path = os.path.join(folder_path, filename)
                pdf_c.drawImage(ImageReader(image_path), x + (CELL_W - QR_SIZE) / 2, y + 10 * mm, width=QR_SIZE, height=QR_SIZE)
                
                # Draw Name
                display_name = name if len(name) <= 22 else name[:20] + "…"
                pdf_c.setFont("Helvetica-Bold", 7)
                pdf_c.setFillColor(colors.black)
                pdf_c.drawCentredString(x + CELL_W / 2, y + 6 * mm, display_name)
                
                # Draw Roll Number
                pdf_c.setFont("Helvetica", 6.5)
                pdf_c.setFillColor(colors.grey)
                pdf_c.drawCentredString(x + CELL_W / 2, y + 2 * mm, roll)
                
                idx += 1
                
        # Draw footer
        pdf_c.setFont("Helvetica", 7)
        pdf_c.setFillColor(colors.grey)
        pdf_c.drawCentredString(PAGE_W / 2, 6 * mm, f"QR Section: {folder_name.replace('_', ' ')} — Page {pdf_c.getPageNumber()}")
        
        if idx < total:
            pdf_c.showPage()
            
    pdf_c.save()
    print(f"   ✓ Generated PDF sheet successfully.")
    return pdf_path

# Generate PDFs
pdf_1st = generate_pdf_for_folder("1st_Year")
pdf_2nd = generate_pdf_for_folder("2nd_Year")

# Recreate Zip
print(f"\n📦 Recreating ZIP package {OUTPUT_ZIP} with new folders and PDFs...")
with zipfile.ZipFile(OUTPUT_ZIP, "w", zipfile.ZIP_DEFLATED) as zf:
    # Pack 1st_Year PNGs
    if os.path.exists("student_qr_codes/1st_Year"):
        for f in sorted(os.listdir("student_qr_codes/1st_Year")):
            if f.endswith(".png"):
                zf.write(os.path.join("student_qr_codes/1st_Year", f), os.path.join("1st_Year", f))
                
    # Pack 2nd_Year PNGs
    if os.path.exists("student_qr_codes/2nd_Year"):
        for f in sorted(os.listdir("student_qr_codes/2nd_Year")):
            if f.endswith(".png"):
                zf.write(os.path.join("student_qr_codes/2nd_Year", f), os.path.join("2nd_Year", f))
                
    # Pack printable PDFs
    if pdf_1st:
        zf.write(pdf_1st, "1st_Year_printable.pdf")
    if pdf_2nd:
        zf.write(pdf_2nd, "2nd_Year_printable.pdf")

print("🎉 New ZIP generated successfully!")
