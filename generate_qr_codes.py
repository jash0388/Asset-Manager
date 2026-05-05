#!/usr/bin/env python3
"""
Generate QR codes for all hostel students from both Excel sheets.
Each QR code encodes the roll number and student name.
All QR codes are saved as PNG images and zipped into a single downloadable file.
"""

import os
import zipfile
import openpyxl
import qrcode
from qrcode.image.styledpil import StyledPilImage
from PIL import Image, ImageDraw, ImageFont

OUTPUT_DIR = "student_qr_codes"
ZIP_FILE = "SPHN_Student_QR_Codes.zip"


def make_qr(roll_no: str, name: str, year_label: str, out_dir: str):
    """Generate a single QR code PNG with roll number + name encoded."""
    roll_no = roll_no.strip()
    name = name.strip()

    # Data encoded in the QR
    qr_data = f"ROLL:{roll_no}|NAME:{name}"

    qr = qrcode.QRCode(
        version=None,  # auto-size
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=4,
    )
    qr.add_data(qr_data)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGB")

    # Create a label below the QR code
    qr_w, qr_h = qr_img.size
    label_height = 80
    canvas = Image.new("RGB", (qr_w, qr_h + label_height), "white")
    canvas.paste(qr_img, (0, 0))

    draw = ImageDraw.Draw(canvas)
    # Try to use a nice font, fall back to default
    try:
        font = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", 20)
        font_small = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", 16)
    except Exception:
        font = ImageFont.load_default()
        font_small = font

    # Draw roll number (bold-ish, centered)
    roll_bbox = draw.textbbox((0, 0), roll_no, font=font)
    roll_tw = roll_bbox[2] - roll_bbox[0]
    draw.text(((qr_w - roll_tw) / 2, qr_h + 5), roll_no, fill="black", font=font)

    # Draw name (centered, smaller)
    name_bbox = draw.textbbox((0, 0), name, font=font_small)
    name_tw = name_bbox[2] - name_bbox[0]
    draw.text(((qr_w - name_tw) / 2, qr_h + 35), name, fill="black", font=font_small)

    # Save
    safe_roll = roll_no.replace("/", "_").replace(" ", "")
    filename = f"{safe_roll}.png"
    filepath = os.path.join(out_dir, filename)
    canvas.save(filepath)
    return filepath


def main():
    # Create output directories
    second_year_dir = os.path.join(OUTPUT_DIR, "2nd_Year")
    first_year_dir = os.path.join(OUTPUT_DIR, "1st_Year")
    os.makedirs(second_year_dir, exist_ok=True)
    os.makedirs(first_year_dir, exist_ok=True)

    all_files = []
    count = 0

    # ── File 1: Second Year Students ──
    wb1 = openpyxl.load_workbook("SPHN_ Hostel Data.xlsx")
    ws1 = wb1["Sheet1"]
    for row in ws1.iter_rows(min_row=3, max_row=ws1.max_row, values_only=True):
        sno, roll_no, name = row[0], row[1], row[2]
        if roll_no is None or name is None:
            continue
        roll_no = str(roll_no).strip()
        name = str(name).strip()
        if not roll_no or not name:
            continue
        fp = make_qr(roll_no, name, "2nd Year", second_year_dir)
        all_files.append(fp)
        count += 1
        print(f"  ✓ [{count}] {roll_no} - {name}")

    # ── File 2: First Year Students ──
    wb2 = openpyxl.load_workbook("SPHN_Hostel Data1.xlsx")
    ws2 = wb2["Sheet1"]
    for row in ws2.iter_rows(min_row=3, max_row=ws2.max_row, values_only=True):
        roll_no, name = row[0], row[1]
        if roll_no is None or name is None:
            continue
        roll_no = str(roll_no).strip()
        name = str(name).strip()
        if not roll_no or not name:
            continue
        fp = make_qr(roll_no, name, "1st Year", first_year_dir)
        all_files.append(fp)
        count += 1
        print(f"  ✓ [{count}] {roll_no} - {name}")

    # ── Create ZIP ──
    print(f"\n📦 Creating ZIP file: {ZIP_FILE}")
    with zipfile.ZipFile(ZIP_FILE, "w", zipfile.ZIP_DEFLATED) as zf:
        for fp in all_files:
            arcname = fp.replace(OUTPUT_DIR + "/", "")
            zf.write(fp, arcname)

    print(f"\n✅ Done! Generated {count} QR codes.")
    print(f"📁 Individual PNGs: ./{OUTPUT_DIR}/")
    print(f"📦 ZIP download:    ./{ZIP_FILE}")


if __name__ == "__main__":
    main()
